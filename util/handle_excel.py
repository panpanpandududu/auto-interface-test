#处理表格
#encoding=utf-8
import openpyxl   #openpyxl是处理Excel表格的模块
import sys
import os
base_path = os.getcwd()
sys.path.append(base_path)

#获取工作簿，拿到Excel表里面的所有数据，即获取层面上的一个对象
#open_excel = openpyxl.load_workbook(base_path+'/case/imooc.xlsx')

#获取Excel里所有表格的名字
#sheet_names = open_excel.sheetnames

#获取Excel里第一个表格的所有数据
#excel_value = open_excel.active    #此方法也可获取第一个sheet
#excel_value = open_excel[sheet_names[0]]
#print(excel_value)            #<Worksheet "Sheet1"> 

#若要获取具体的单元格，用cell():返回单元格对象
#excel_obj = excel_value.cell(1,4)
#print(excel_obj)              #<Cell 'Sheet1'.D1>

#获取指定单元格的值
#print(excel_obj.value)    #结果为前置条件
#print(excel_value.max_row)  #10


#封装Excel表格的操作
class HandExcel:
    def load_excel(self):
        '''
        加载excel
        '''
        open_value = openpyxl.load_workbook(base_path+'/case/imooc.xlsx')
        return open_value
    def get_sheet_data(self,index=None):
        '''
         加载某一sheet的所有数据
        '''
        sheet_names = self.load_excel().sheetnames  #获取所有的sheet
        if index ==None:
            index = 0
        data = self.load_excel()[sheet_names[index]] #获取某一sheet
        return data
    
    def get_cell_value(self,row,cols):
        '''
        获取某一单元格的内容
        '''
        value = self.get_sheet_data().cell(row=row,column=cols).value

        return value
    
    def get_rows(self):
        '''
        获取行数
        '''
        row = self.get_sheet_data().max_row
        return row

    def get_rows_data(self,row):
        '''
        获取某一行内容
        '''
        row_list =[] 
        for i in self.get_sheet_data()[row]:    #获取行对象：self.get_sheet_data()[row]
            row_list.append(i.value)            #获取行内容：self.get_sheet_data()[row].value
        return row_list

    def get_columns_data(self,key=None):
        '''
        获取某一列的内容
        '''
        columns_list = []
        if key ==None:
            key = 'A'
        for i in self.get_sheet_data()[key]:
            columns_list.append(i.value)
        return columns_list
    
    def get_rows_number(self,case_id):
        '''
        数据依赖中根据caseid获取某一行号
        '''
        num = 1
        for i in self.get_columns_data():
            if i == case_id:
                return num
            num = num+1
        return num
        
    
    def excel_write_data(self,rows,cols,value):
        '''写入数据，即执行结果回写'''
        wb = self.load_excel()  #获取表格
        wr = wb.active  #激活表格，然后可写入 
        wr.cell(rows,cols,value)   #写入数据
        wb.save(base_path+"/case/imooc.xlsx")  #保存数据

    
    def get_excel_data(self):
        '''
        获取excel所有数据     #方便后面数据驱动拿数据
        '''
        data_list = []
        for i in range(self.get_rows()):
            data_list.append(self.get_rows_data(i+2))
        return data_list




 

#实例化类
excel_data = HandExcel()

#执行程序
# if __name__ == "__main__":
#     print(excel_data.get_excel_data())
#     '程序运行'
    # data = excel_data.get_cell_value(2,7)
    # rows_data = excel_data.get_rows_data(3)
    # print(rows_data)
    # data_rows = excel_data.get_rows()
    #print(data_rows)  #10
    # print(data)   #post

